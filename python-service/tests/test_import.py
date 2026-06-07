"""导入功能单元测试"""
import io
import pytest
from services.import_service import parse_excel_to_docs


class TestParseExcel:
    def test_parse_simple_excel(self):
        """解析基本 Excel"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["_id", "title", "rating"])
        ws.append(["fb_0001", "测试标题", 5])
        ws.append(["fb_0002", "另一标题", 3])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        docs = parse_excel_to_docs(buf.read())
        assert len(docs) == 2
        assert docs[0]["_id"] == "fb_0001"
        assert docs[0]["title"] == "测试标题"
        assert docs[0]["rating"] == 5

    def test_parse_skips_empty_header(self):
        """跳过空的表头列"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["_id", None, "title"])
        ws.append(["fb_0001", "ignored", "测试"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        docs = parse_excel_to_docs(buf.read())
        assert len(docs) == 1
        assert "_id" in docs[0]
        assert "title" in docs[0]
        assert None not in docs[0]

    def test_parse_skips_empty_rows(self):
        """跳过全空行"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["_id", "title"])
        ws.append([None, None])  # 空行
        ws.append(["fb_0001", "测试"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        docs = parse_excel_to_docs(buf.read())
        assert len(docs) == 1

    def test_parse_empty_file(self):
        """空文件返回空列表"""
        import openpyxl
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        docs = parse_excel_to_docs(buf.read())
        assert docs == []
